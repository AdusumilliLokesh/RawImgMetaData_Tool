import os
from openpyxl import Workbook
import pandas as pd
import json

# open the config file for static data
with open('config.json', 'r') as file:
    data = json.load(file)

# Convert keys to variables (adds them to global namespace)
for key, value in data.items():
    globals()[key] = value

# input file setup
input_file_path = os.path.join(input_path, input_file_name)

# output file setup
output_file_path = os.path.join(output_path, output_file_name)

# Required columns expected in the input file
required_columns = {"Date", "Total_Pages", "vol", "issue"}

try:
    # read raw data
    df = pd.read_excel(input_file_path)

    # check for required columns
    missing_cols = required_columns - set(df.columns)
    if missing_cols:
        raise ValueError(f"Missing required column(s): {', '.join(missing_cols)}")

except Exception as e:
    print(f"Error reading input file: {e}")
    print("Please ensure the input Excel file exactly matches the expected column names: Date, Total_Pages, vol, and issue.")
    exit(1)
# read raw data
df = pd.read_excel(input_file_path)

# Create a workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "MetaData"

# Static headers for the output file
headers = [
    "Title", "Standard_Title", "LCCN", "Date", "Date-Numeric", "Volume_Number",
    "Issue_Number", "Edition_Order", "Page_Sequence_Number", "Page_Physical_Description",
    "Frame_ID", "Issue_ID", "Reproduction_Agency", "Reproduction_Agency_Code",
    "Reproduction_Note", "Physical_Location", "Physical_Location_Code", "Batch_Name",
    "Reel_Number", "Reel_Sequence_Number", "Digital_Filename"
]

ws.append(headers) #append headers to o/p file
# set-up config variables
row_count = 1
df['Date_in_words'] = df['Date'].dt.strftime('%B %-d, %Y')
df['Date '] = pd.to_datetime(df['Date']).dt.date
df['Formatted_Title'] = Standard_Title + ' [' + city + ' ' + state + ', ' + df['Date_in_words'] + ']'
df['Digital_Filename'] = Reel_Number + "-"

# Add headers and data
for i, row in df.iterrows():
    for j in range(1,df.loc[i,"Total_Pages"]+1):
        print(df.loc[i,"Date "],j)
        row_count = row_count + 1
        ws.cell(row=row_count, column=1, value=df.loc[i, "Formatted_Title"])
        ws.cell(row=row_count, column=2, value=Standard_Title)
        ws.cell(row=row_count, column=3, value=LCCN)
        ws.cell(row=row_count, column=4, value=df.loc[i, "Date_in_words"])
        ws.cell(row=row_count, column=5, value=df.loc[i,"Date"].date())
        print(df.loc[i,"Date"].date())
        ws.cell(row=row_count, column=6, value=df.loc[i, "vol"])
        ws.cell(row=row_count, column=7, value=df.loc[i, "issue"])
        ws.cell(row=row_count, column=8, value=Edition_Order)
        ws.cell(row=row_count, column=9, value=j)
        ws.cell(row=row_count, column=10, value=Page_Physical_Description)
        ws.cell(row=row_count, column=11, value=Reel_Number+'_'+str(df.loc[i,"Date"])+'_'+Edition_Order+'_'+str(j).zfill(len(str(df.loc[i,"Total_Pages"]))))
        ws.cell(row=row_count, column=12, value=Reel_Number + '_' + str(df.loc[i, "Date"]) + '_' + Edition_Order)
        ws.cell(row=row_count, column=13, value=Reproduction_Agency)
        ws.cell(row=row_count, column=14, value=Reproduction_Agency_Code)
        ws.cell(row=row_count, column=15, value=Reproduction_Note)
        ws.cell(row=row_count, column=16, value=Physical_Location)
        ws.cell(row=row_count, column=17, value=Physical_Location_Code)
        ws.cell(row=row_count, column=18, value=Batch_Name)
        ws.cell(row=row_count, column=19, value=Reel_Number)
        ws.cell(row=row_count, column=20, value=row_count-1)
        ws.cell(row=row_count, column=21, value=df.loc[i, "Digital_Filename"]+ (str(row_count-1)).zfill(4))

# Save the file
wb.save(output_file_path)
print(f"Excel file saved successfully to: {output_file_path}")
